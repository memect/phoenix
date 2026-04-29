"""从 workspace 执行提取并输出 Excel 报告

用法:
    uv run python scripts/run_extract_to_excel.py <workspace_path> [--output <output.xlsx>]

示例:
    uv run python scripts/run_extract_to_excel.py local/workspaces/schema_audit_opinion
    uv run python scripts/run_extract_to_excel.py local/workspaces/schema_audit_opinion -o /tmp/result.xlsx
"""

import argparse
import asyncio
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from extract_dev.api import extract_from_docjson


def load_doc_ids(data_dir: Path) -> list[str]:
    train_file = data_dir / "standard_for_evaluate" / "train.json"
    if not train_file.exists():
        raise FileNotFoundError(f"找不到 {train_file}")
    entries = json.loads(train_file.read_text())
    return [e["document_id"] for e in entries]


async def run_extraction(workspace: Path) -> dict[str, dict]:
    """对 workspace 下所有文档执行提取"""
    data_dir = workspace / ".extract-dev" / "data"
    doc_ids = load_doc_ids(data_dir)
    print(f"共 {len(doc_ids)} 个文档")

    results = {}
    for i, doc_id in enumerate(doc_ids, 1):
        docjson_path = data_dir / "docjson" / f"{doc_id.replace('-', '')}.json"
        docjson = json.loads(docjson_path.read_text())
        r = await extract_from_docjson(docjson, workspace=str(workspace))
        results[doc_id] = r if isinstance(r, dict) else {}
        print(f"[{i}/{len(doc_ids)}] {doc_id[:8]}... done")

    return results


def write_excel(results: dict[str, dict], output_path: Path) -> None:
    """将提取结果写入 Excel"""
    if not results:
        print("无结果可写入")
        return

    # 从第一个结果推断字段列表
    field_keys = []
    for data in results.values():
        if data:
            field_keys = list(data.keys())
            break

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "提取结果"

    # 样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    cell_font = Font(size=10)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")

    # 写表头
    headers = ["文档ID"] + field_keys
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # 写数据行
    for row_idx, (doc_id, data) in enumerate(results.items(), 2):
        cell = ws.cell(row=row_idx, column=1, value=doc_id)
        cell.font = cell_font
        cell.border = border
        cell.alignment = center
        for col_idx, key in enumerate(field_keys, 2):
            val = data.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val else "-")
            cell.font = cell_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 自动列宽（简单估算）
    ws.column_dimensions["A"].width = 16
    for col_idx, key in enumerate(field_keys, 2):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        max_len = max(len(key), max((len(str(d.get(key, ""))) for d in results.values()), default=0))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

    wb.save(output_path)
    print(f"Excel 已保存: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="从 workspace 执行提取并输出 Excel")
    parser.add_argument("workspace", type=Path, help="workspace 路径")
    parser.add_argument("-o", "--output", type=Path, default=None, help="输出 Excel 路径")
    args = parser.parse_args()

    workspace = args.workspace.resolve()
    if not (workspace / "program.py").exists():
        raise FileNotFoundError(f"workspace 中没有 program.py: {workspace}")

    output = args.output or (workspace / "extract_results.xlsx")

    results = asyncio.run(run_extraction(workspace))
    write_excel(results, output)


if __name__ == "__main__":
    main()
